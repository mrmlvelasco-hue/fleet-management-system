"""Tests for the three operational features:
  * vehicle bulk import from the fill-in Excel template
  * `flask reset transactions` (clears transactions, keeps master data)
  * the dashboard due-lists excluding work that's already been raised
"""
from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.cli import _seed_transaction_types, _reset_transactions
from app.modules.master_data.reference.service import (
    VehicleTypeService, MaintenanceTypeService)
from app.modules.master_data.org.service import BranchService
from app.modules.master_data.vehicle.service import VehicleService
from app.modules.master_data.vehicle.import_service import (
    build_template, import_vehicles)
from app.modules.transactions.maintenance_order.service import (
    MaintenanceOrderService)


@pytest.fixture()
def env(db):
    vt = VehicleTypeService().create(code="TRK", name="Truck",
                                     category="HEAVY")
    branch = BranchService().create(code="MNL", name="Manila")
    return vt, branch


def _filled_template(rows):
    """Build a workbook from the real template, example row removed,
    then append the given rows."""
    wb = load_workbook(BytesIO(build_template()))
    ws = wb["Vehicles"]
    ws.delete_rows(2, 2)  # example row + the 'delete me' note
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Vehicle import ──────────────────────────────────────────────────────────

def test_template_has_expected_sheets_and_headers(db, env):
    wb = load_workbook(BytesIO(build_template()))
    assert wb.sheetnames == ["Vehicles", "Reference"]
    headers = [c.value for c in wb["Vehicles"][1]]
    for required in ("vehicle_type_code", "branch_code", "brand", "model",
                     "year"):
        assert required in headers


def test_template_reference_sheet_lists_this_installs_codes(db, env):
    """The template must carry the ACTUAL valid codes, so whoever fills
    it in doesn't have to guess or look them up elsewhere."""
    wb = load_workbook(BytesIO(build_template()))
    text = "\n".join(
        str(c.value) for row in wb["Reference"].iter_rows() for c in row
        if c.value is not None)
    assert "TRK" in text     # the seeded vehicle type code
    assert "MNL" in text     # the seeded branch code


def test_dry_run_reports_without_writing_anything(db, env):
    vt, branch = env
    from app.modules.master_data.vehicle.models import Vehicle
    before = Vehicle.query.count()
    buf = _filled_template([
        ["IMP-001", "PLATE-1", "TRK", "MNL", "Toyota", "Vios", 2021,
         "E", "Silver", "ENG1", "CHS1", "GASOLINE", "MANUAL", 12000,
         "2021-03-01", 850000, "FAR-9", "CR-9", "ACTIVE"],
    ])
    result = import_vehicles(buf, dry_run=True)
    assert result["created"] == 1
    assert result["skipped"] == 0
    assert Vehicle.query.count() == before  # nothing written


def test_real_import_creates_the_vehicle_with_parsed_fields(db, env):
    buf = _filled_template([
        ["IMP-001", "PLATE-1", "TRK", "MNL", "Toyota", "Vios", 2021,
         "E", "Silver", "ENG1", "CHS1", "GASOLINE", "MANUAL", 12000,
         "2021-03-01", 850000, "FAR-9", "CR-9", "ACTIVE"],
    ])
    result = import_vehicles(buf, dry_run=False)
    assert result["created"] == 1

    from app.modules.master_data.vehicle.models import Vehicle
    v = Vehicle.query.filter_by(conduction_number="IMP-001").first()
    assert v is not None
    assert (v.brand, v.model, v.year) == ("Toyota", "Vios", 2021)
    assert v.current_odometer == 12000
    assert v.acquisition_date == date(2021, 3, 1)


def test_bad_rows_are_skipped_and_reported_while_good_rows_import(db, env):
    """A migration file must not be all-or-nothing: one bad row should
    not block the other 399."""
    buf = _filled_template([
        ["IMP-OK", "PLATE-OK", "TRK", "MNL", "Toyota", "Vios", 2021,
         "", "", "", "", "", "", "", "", "", "", "", ""],
        # no conduction AND no plate
        ["", "", "TRK", "MNL", "Ford", "Ranger", 2020,
         "", "", "", "", "", "", "", "", "", "", "", ""],
        # unknown vehicle type code
        ["IMP-BAD2", "PLATE-BAD2", "NOPE", "MNL", "Nissan", "Navara", 2019,
         "", "", "", "", "", "", "", "", "", "", "", ""],
    ])
    result = import_vehicles(buf, dry_run=False)
    assert result["created"] == 1
    assert result["skipped"] == 2
    problems = " ".join(
        p for e in result["errors"] for p in e["problems"])
    assert "conduction_number or a plate_number" in problems
    assert "unknown vehicle_type_code" in problems

    from app.modules.master_data.vehicle.models import Vehicle
    assert Vehicle.query.filter_by(conduction_number="IMP-OK").first()


def test_duplicate_plate_is_reported_not_crashed(db, env):
    vt, branch = env
    VehicleService().create(vehicle_type_id=vt.id, brand="Toyota",
                            model="Vios", year=2020, branch_id=branch.id,
                            plate_number="DUPE-1")
    buf = _filled_template([
        ["IMP-D", "DUPE-1", "TRK", "MNL", "Toyota", "Vios", 2021,
         "", "", "", "", "", "", "", "", "", "", "", ""],
    ])
    result = import_vehicles(buf, dry_run=False)
    assert result["skipped"] == 1
    assert "already exists" in result["errors"][0]["problems"][0]


def test_leftover_example_row_is_ignored_not_imported(db, env):
    """People routinely forget to delete the template's example row --
    importing it as a real vehicle would be worse than ignoring it."""
    wb = load_workbook(BytesIO(build_template()))  # example row intact
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    result = import_vehicles(buf, dry_run=True)
    assert result["created"] == 0

    from app.modules.master_data.vehicle.models import Vehicle
    assert Vehicle.query.filter_by(
        conduction_number="EXAMPLE-001").first() is None


# ── Transaction reset ───────────────────────────────────────────────────────

def test_reset_clears_transactions_but_keeps_master_data(db, env):
    vt, branch = env
    _seed_transaction_types()
    mt = MaintenanceTypeService().create(code="PMS-R", name="PMS",
                                         category="PREVENTIVE")
    vehicle = VehicleService().create(
        vehicle_type_id=vt.id, brand="Toyota", model="Hilux", year=2022,
        branch_id=branch.id, conduction_number="RST-1")
    MaintenanceOrderService().create(
        vehicle_id=vehicle.id, maintenance_type_id=mt.id,
        scheduled_date=date.today(), user=None)

    from app.modules.transactions.maintenance_order.models import (
        MaintenanceOrder)
    from app.modules.master_data.vehicle.models import Vehicle
    assert MaintenanceOrder.query.count() >= 1

    _reset_transactions()

    assert MaintenanceOrder.query.count() == 0     # transactions gone
    assert Vehicle.query.count() >= 1              # master data kept
    assert Vehicle.query.filter_by(conduction_number="RST-1").first()


# ── Due lists exclude already-raised work ───────────────────────────────────

def test_due_maintenance_excludes_vehicle_with_an_open_order(db, env):
    """A vehicle whose PM has already been raised must drop off the
    'Vehicles Due for Maintenance' list -- otherwise the dashboard keeps
    offering an action that creating would refuse."""
    vt, branch = env
    from app.modules.maintenance_config.service import PMScheduleService
    from app.core.maintenance.due_calculation_service import (
        PMDueCalculationService)
    mt = MaintenanceTypeService().create(code="PMS-DUE", name="PMS",
                                         category="PREVENTIVE")
    PMScheduleService().create(
        vehicle_type_id=vt.id, maintenance_type_id=mt.id,
        trigger_mode="KM", interval_km=5000)
    vehicle = VehicleService().create(
        vehicle_type_id=vt.id, brand="Toyota", model="Hilux", year=2022,
        branch_id=branch.id, conduction_number="DUE-1")
    vehicle.current_odometer = 99000
    db.session.commit()

    svc = PMDueCalculationService()
    before = [d for d in svc.get_all_due_vehicles()
             if d["vehicle"].id == vehicle.id]
    assert before, "vehicle should be due before any order is raised"

    MaintenanceOrderService().create(
        vehicle_id=vehicle.id, maintenance_type_id=mt.id,
        scheduled_date=date.today(), user=None)

    after = [d for d in svc.get_all_due_vehicles()
            if d["vehicle"].id == vehicle.id]
    assert not after, "vehicle with an open PM order must not still be listed"

    # ...but the raw calculation can still see it when asked explicitly.
    raw = [d for d in svc.get_all_due_vehicles(exclude_with_open_order=False)
          if d["vehicle"].id == vehicle.id]
    assert raw
