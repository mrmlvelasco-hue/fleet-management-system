"""Migration: import PM_Task_List.xlsx ('Sheet1') into PMSchedule (+
PMScopeTemplate/PMScopeItem) — a separate, purpose-built importer for
this specific file, NOT a modification of import_vems_pms.py (which
targets a different source file/sheet layout, 'VEMS_Masterdata_for_
vehicle.xlsx''s 'PMS' sheet, with its own passing tests already
depending on that exact positional structure).

IMPORTANT — how this differs from the other importer's approach:
import_vems_pms.py infers a package's true recurring interval from the
PERIODICITY of repeated identical Scope-text occurrences, because that
source file pre-expands one clean recurring package across many rows
sharing byte-identical checklist text. Checked whether the same pattern
holds here (see the investigation this file's migration prompted) and
it does NOT: the overwhelming majority of rows in THIS file have a
genuinely UNIQUE Scope text per row (Ford Escape's 40 PM rows produced
only 3 scope-text collisions out of ~37 distinct texts, with irregular,
non-uniform gaps between the few that did repeat) -- meaning each row
here really is its own distinct package, not a redundantly pre-expanded
copy of a smaller recurring set. Inferring periodicity from repetition
would therefore be actively WRONG for this file: it silently produced
multiplied, incorrect intervals (e.g. "10,000" instead of "5,000") that
didn't match the row's own WorkDescription text at all when first
tried against this exact data.

The correct, directly-verifiable model for this file: each row's own
KM Reading / Calendar / Hourly code already states its actual, real
interval outright -- no inference required. Exact duplicate rows
(identical Scope + WorkDescription + all three codes) are collapsed,
but that's it.

Three real gaps fixed vs. how this kind of file was previously migrated:

  1. WorkDescription (the per-package, pm2-pm9-tokenized work-order
     description template, e.g. "First 1,000 km servicing of pm2 pm3
     with Plate no. pm4...") was never captured anywhere — only the
     generic category-level Description. Now captured into
     PMSchedule.work_description_template.

  2. The frequency codes (e.g. "10KMS", "4.5MTH") are decoded via a
     corrected, validated reference table (vems_frequency_reference.py)
     instead of trusting the source lookup's Val/ValDesc columns as-is
     (several have real data-entry errors) or fragile regex-parsing the
     code string (which can't handle decimals or Quarter/Week/Hour
     units at all).

  3. Month/Year/Week/Quarter intervals are converted into interval_days
     (30/365/7/90 days respectively), as requested.

Reads columns by HEADER NAME, not positional index.

"Vehicle Registration" rows are excluded — LTO annual renewal, already
covered by the dedicated Vehicle Registration transaction module.
"""
import os
import re
import sys
from collections import defaultdict

# Make sure THIS project's app/ package is what gets imported, regardless
# of the current working directory or PYTHONPATH -- and regardless of
# whether some unrelated pip-installed package also happens to be named
# "app" (this really happened during testing: an ImportStringError
# traceback showed Python resolving `app` to
# .venv\Lib\site-packages\app\__init__.py instead of this project's own
# app/ folder). Inserting the real project root at sys.path[0] (the
# directory ABOVE this scripts/ folder) makes the local package win
# every time, and removes any dependency on how the script happens to be
# invoked (bare `python scripts\...py`, a different cwd, no PYTHONPATH
# set at all, etc.) -- the previous version of this script assumed the
# caller had already arranged that, which silently wasn't true the
# moment it was actually run standalone from PowerShell.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# The `flask` CLI (flask db upgrade, flask seed all, ...) auto-loads a
# .env file for you -- that's WHY `flask db upgrade` correctly reached
# the real MySQL database and added the new columns there. A raw
# `python scripts\...py` invocation does NOT go through that CLI
# bootstrapping at all, so DATABASE_URL/DB_* env vars from .env were
# simply never set in that process -- config.py's own fallback then
# silently defaulted to a brand-new, never-migrated
# sqlite:///fms_dev.db, which is exactly the "no such column:
# pm_schedules.interval_hours" error: that column really does exist on
# the real MySQL database (confirmed via SHOW CREATE TABLE), just not on
# this accidental, empty, unrelated SQLite file this script ended up
# talking to instead. Loading .env here ourselves makes this script
# connect to the exact same database `flask` commands do, every time.
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

import openpyxl

from app.extensions import db
from app.modules.maintenance_config.service import (
    PMScheduleService, PMScopeTemplateService)
from app.modules.master_data.reference.service import MaintenanceTypeService
from app.modules.master_data.vehicle_brand.service import (
    VehicleBrandService, VehicleModelService)
from vems_frequency_reference import (
    to_interval_km, to_interval_days, to_interval_hours)

EXCLUDED_CATEGORIES = {"Vehicle Registration"}

CATEGORY_TO_MTYPE = {
    "Vehicle Preventive Maintenance": ("PMS-PREV", "Preventive Maintenance Service"),
    "Tire Replacement": ("PMS-TIRE", "Tire Replacement"),
    "Battery Replacement": ("PMS-BATT", "Battery Replacement"),
    "Aircon Servicing": ("PMS-AIRCON", "Aircon Servicing"),
}


def _split_scope_into_items(scope_text: str) -> list:
    """Splits VEMS's numbered checklist text ('1. Foo  2. Bar') into
    individual activity strings."""
    if not scope_text:
        return []
    parts = re.split(r"\d{1,2}\.\s+", scope_text)
    return [p.strip() for p in parts if p.strip()]


def _hours_from_hourly_column(value):
    """The Hourly column holds either a real frequency code ('500HRS')
    or, in some rows, a bare number ('500') meaning that many hours
    directly. Tries the code table first, then falls back to treating a
    plain digit string as the raw hour count."""
    hrs = to_interval_hours(value)
    if hrs is not None:
        return hrs
    s = str(value).strip() if value is not None else ""
    if s.isdigit() and int(s) > 0:
        return int(s)
    return None


def _get_or_create_maintenance_type(code: str, name: str, cache: dict):
    if code in cache:
        return cache[code]
    from app.modules.master_data.reference.models import MaintenanceType
    existing = MaintenanceType.query.filter_by(code=code).first()
    if existing:
        cache[code] = existing
        return existing
    created = MaintenanceTypeService().create(code=code, name=name,
                                               category="PM")
    cache[code] = created
    return created


def _resolve_brand_model(make: str, model: str, brand_cache: dict):
    key = (make or "").strip().lower()
    if key not in brand_cache:
        brand_cache[key] = VehicleBrandService().get_by_name(make)
    brand = brand_cache[key]
    if not brand:
        return None, None
    model_obj = VehicleModelService().get_by_name_and_brand(model, brand.id)
    return brand.id, (model_obj.id if model_obj else None)


def import_pm_task_list(xlsx_path: str, dry_run: bool = True,
                        limit_groups: int = None) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(header_row)}
    required = ["Make", "Model", "Description", "Task_CD", "WorkDescription",
               "Scope", "KM Reading", "Calendar", "Hourly", "PMDescription",
               "Sort"]
    missing = [h for h in required if h not in idx]
    if missing:
        raise ValueError(f"Expected column(s) not found in {xlsx_path}: "
                         f"{missing}. Found: {header_row}")

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    groups = defaultdict(list)
    for r in rows:
        groups[r[idx["Task_CD"]]].append(r)

    mtype_cache, brand_cache = {}, {}
    stats = {
        "groups_processed": 0, "groups_skipped_excluded": 0,
        "groups_skipped_no_category": 0, "packages_created": 0,
        "scope_items_created": 0, "unrecognized_frequency_codes": set(),
        "hour_based_packages": 0, "duplicate_rows_collapsed": 0,
        "samples": [],
    }

    group_items = list(groups.items())
    if limit_groups:
        group_items = group_items[:limit_groups]

    for task_cd, group_rows in group_items:
        first = group_rows[0]
        category = first[idx["PMDescription"]]
        if category in EXCLUDED_CATEGORIES:
            stats["groups_skipped_excluded"] += 1
            continue
        if category not in CATEGORY_TO_MTYPE:
            stats["groups_skipped_no_category"] += 1
            continue

        make = first[idx["Make"]]
        model = first[idx["Model"]]
        task_description = first[idx["Description"]]

        # Collapse only EXACT duplicate rows (identical Scope +
        # WorkDescription + all three frequency codes) -- everything
        # else is treated as its own distinct package, per the
        # investigation in this module's docstring.
        seen = {}
        for r in group_rows:
            key = (r[idx["Scope"]], r[idx["WorkDescription"]],
                  r[idx["KM Reading"]], r[idx["Calendar"]], r[idx["Hourly"]])
            sort_val = int(r[idx["Sort"]]) if r[idx["Sort"]] not in (None, "") else 1
            if key in seen:
                stats["duplicate_rows_collapsed"] += 1
                if sort_val < seen[key][0]:
                    seen[key] = (sort_val, r)
            else:
                seen[key] = (sort_val, r)

        packages = sorted(seen.values(), key=lambda p: p[0])  # by Sort

        mtype_code, mtype_name = CATEGORY_TO_MTYPE[category]
        mtype = _get_or_create_maintenance_type(mtype_code, mtype_name, mtype_cache) \
            if not dry_run else mtype_cache.setdefault(mtype_code, None)

        brand_id, model_id = (None, None)
        if not dry_run:
            brand_id, model_id = _resolve_brand_model(make, model, brand_cache)

        for seq_pos, (_sort_val, row) in enumerate(packages, start=1):
            scope_text = row[idx["Scope"]]
            work_desc = row[idx["WorkDescription"]]
            km_code = row[idx["KM Reading"]]
            cal_code = row[idx["Calendar"]]
            hourly_raw = row[idx["Hourly"]]

            interval_km = to_interval_km(km_code)
            interval_days = to_interval_days(cal_code)
            interval_hours = _hours_from_hourly_column(hourly_raw)

            for code in (km_code, cal_code):
                if code not in (None, 0, "0") and to_interval_km(code) is None \
                        and to_interval_days(code) is None:
                    stats["unrecognized_frequency_codes"].add(str(code))
            if interval_hours:
                stats["hour_based_packages"] += 1

            trigger_mode = ("HYBRID" if interval_km and interval_days
                           else "KM" if interval_km else "CALENDAR")
            if not interval_km and not interval_days:
                continue  # nothing usable to schedule on

            activity_texts = _split_scope_into_items(scope_text)
            if len(stats["samples"]) < 10:
                stats["samples"].append({
                    "task_cd": task_cd, "make": make, "model": model,
                    "category": category, "sequence": seq_pos,
                    "interval_km": interval_km, "interval_days": interval_days,
                    "interval_hours": interval_hours,
                    "trigger_mode": trigger_mode,
                    "activity_count": len(activity_texts),
                    "work_description": (str(work_desc)[:80] if work_desc else ""),
                })

            if not dry_run:
                sched = PMScheduleService().create(
                    maintenance_type_id=mtype.id, trigger_mode=trigger_mode,
                    vehicle_make=str(make) if brand_id is None else None,
                    vehicle_model=str(model) if brand_id is None else None,
                    vehicle_brand_id=brand_id, vehicle_model_id=model_id,
                    profile_code=str(task_cd),
                    profile_description=str(task_description),
                    sequence_position=seq_pos,
                    interval_km=interval_km, interval_days=interval_days,
                    interval_hours=interval_hours,
                    work_description_template=(str(work_desc) if work_desc else None),
                    priority="MEDIUM")
                if activity_texts:
                    items = [{
                        "activity_code": f"{task_cd}-{i+1:03d}",
                        "activity_description": text[:255],
                        "sort_order": i,
                    } for i, text in enumerate(activity_texts)]
                    PMScopeTemplateService().create(
                        maintenance_type_id=mtype.id,
                        name=f"{task_description} - Package {seq_pos}"[:120],
                        description=str(task_description),
                        pm_schedule_id=sched.id, items=items)
            stats["scope_items_created"] += len(activity_texts)
            stats["packages_created"] += 1

        stats["groups_processed"] += 1

    if not dry_run:
        db.session.commit()

    stats["unrecognized_frequency_codes"] = sorted(
        stats["unrecognized_frequency_codes"])
    return stats


def reset_pm_data() -> dict:
    """Wipes ALL PMSchedule/PMScopeTemplate/PMScopeItem rows -- for
    starting a clean re-migration, per the explicit request to "delete
    all the template in the table and re-migrate again" (e.g. after
    fixing a database character-set issue that corrupted a previous
    import).

    Real bug fixed here: the first version of this function only
    cleared PMScopeItem -> PMScopeTemplate -> PMSchedule (the three
    tables actually being reset), but missed that OTHER tables
    reference these by foreign key too:
      - Vehicle.pm_schedule_id ("Assigned PM Template" on Vehicle Master)
      - MaintenanceOrder.pm_schedule_id
      - MaintenanceOrder.scope_template_id
    On MySQL (which enforces FK constraints by default, unlike SQLite's
    more lenient default), deleting a referenced pm_scope_templates row
    while a real MaintenanceOrder still points at it via
    scope_template_id fails outright with "Cannot delete or update a
    parent row: a foreign key constraint fails" -- exactly the reported
    error. This does NOT delete those Vehicle/MaintenanceOrder rows (no
    transaction history is lost) -- it only detaches their reference to
    the specific template row being replaced, which is precisely what
    "wipe the PM catalog and re-migrate fresh" means: the actual
    completed work orders and their own recorded checklist data stay
    exactly as they were; only the link to the OLD catalog entry (which
    is about to be replaced by a freshly re-imported one) is cleared.

    This is intentionally its own explicit opt-in step (--reset flag),
    never automatic, since it touches real Vehicle/MaintenanceOrder rows
    (only to null a reference, not delete them) in addition to wiping
    all PM schedule data."""
    from app.modules.maintenance_config.models import (
        PMSchedule, PMScopeTemplate, PMScopeItem)
    from app.modules.master_data.vehicle.models import Vehicle
    from app.modules.transactions.maintenance_order.models import (
        MaintenanceOrder)

    counts = {
        "pm_scope_items": PMScopeItem.query.count(),
        "pm_scope_templates": PMScopeTemplate.query.count(),
        "pm_schedules": PMSchedule.query.count(),
        "vehicles_unlinked": Vehicle.query.filter(
            Vehicle.pm_schedule_id.isnot(None)).count(),
        "maintenance_orders_unlinked": MaintenanceOrder.query.filter(
            db.or_(MaintenanceOrder.pm_schedule_id.isnot(None),
                  MaintenanceOrder.scope_template_id.isnot(None))).count(),
    }

    # Detach external references FIRST -- these updates, not deletes, so
    # no Vehicle or MaintenanceOrder row is touched beyond this one
    # column each.
    Vehicle.query.filter(Vehicle.pm_schedule_id.isnot(None)).update(
        {"pm_schedule_id": None})
    MaintenanceOrder.query.filter(
        MaintenanceOrder.pm_schedule_id.isnot(None)).update(
        {"pm_schedule_id": None})
    MaintenanceOrder.query.filter(
        MaintenanceOrder.scope_template_id.isnot(None)).update(
        {"scope_template_id": None})
    db.session.commit()

    # Now the three PM tables themselves, children before parents.
    PMScopeItem.query.delete()
    PMScopeTemplate.query.delete()
    PMSchedule.query.delete()
    db.session.commit()
    return counts


if __name__ == "__main__":
    import sys
    from app import create_app
    path = sys.argv[1] if len(sys.argv) > 1 else "PM_Task_List.xlsx"
    dry = "--dry-run" in sys.argv
    do_reset = "--reset" in sys.argv
    app = create_app()
    with app.app_context():
        if do_reset:
            if dry:
                print("--reset ignored during --dry-run (nothing is ever "
                     "deleted in a dry run).")
            else:
                deleted = reset_pm_data()
                print(f"Reset: deleted {deleted['pm_schedules']} PM "
                     f"schedules, {deleted['pm_scope_templates']} scope "
                     f"templates, {deleted['pm_scope_items']} scope items. "
                     f"Detached (not deleted) the old template reference "
                     f"from {deleted['vehicles_unlinked']} vehicles and "
                     f"{deleted['maintenance_orders_unlinked']} maintenance "
                     f"orders that pointed at the old catalog.")
        result = import_pm_task_list(path, dry_run=dry)
        for k, v in result.items():
            if k != "samples":
                print(k, ":", v)
        print("\nSample packages:")
        for s in result["samples"]:
            print(s)
