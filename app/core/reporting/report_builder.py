"""Custom Report Builder — curated data sources and safe query building.

Deliberately NOT a "paste your own SELECT" box. Everything a person can
choose comes from the whitelist below, and the SQL is constructed by
SQLAlchemy from those choices, so:

  * a report can only ever read tables and columns that appear here --
    no reaching into users, password hashes, or the audit trail;
  * values are always bound parameters, so a filter value can never
    alter the query's structure (there is no string concatenation of
    user input into SQL anywhere in this module);
  * each source declares the permission its data already requires, so a
    custom report cannot be used to see data the person is not allowed
    to see through the normal screens.

Adding a new source or column is an explicit, reviewable code change
rather than something a user can do by typing a table name.
"""
from datetime import date, datetime

from sqlalchemy import and_, or_

from app.extensions import db


class Field:
    """One selectable column."""

    def __init__(self, key, label, attr, kind="string", join=None):
        self.key = key          # stable identifier stored in saved reports
        self.label = label      # what the person sees
        self.attr = attr        # dotted path from the base model
        self.kind = kind        # string | number | date | boolean
        self.join = join        # which declared join this field needs


class Join:
    """A pre-declared, allowed relationship. Users pick fields, never
    joins directly -- the builder works out which joins are needed."""

    def __init__(self, key, relationship, label):
        self.key = key
        self.relationship = relationship
        self.label = label


class DataSource:
    def __init__(self, key, label, model_path, permission, joins, fields,
                 description=""):
        self.key = key
        self.label = label
        self.model_path = model_path
        self.permission = permission
        self.description = description
        self.joins = {j.key: j for j in joins}
        self.fields = {f.key: f for f in fields}

    @property
    def model(self):
        import importlib
        module_path, class_name = self.model_path.rsplit(".", 1)
        return getattr(importlib.import_module(module_path), class_name)


DATA_SOURCES = {
    "vehicles": DataSource(
        key="vehicles", label="Vehicles",
        model_path="app.modules.master_data.vehicle.models.Vehicle",
        permission="vehicle.view",
        description="Fleet master data, with branch and vehicle type.",
        joins=[
            Join("branch", "branch", "Branch"),
            Join("vehicle_type", "vehicle_type", "Vehicle Type"),
            Join("assigned_driver", "assigned_driver", "Assigned Driver"),
        ],
        fields=[
            Field("plate_number", "Plate No.", "plate_number"),
            Field("conduction_number", "Conduction No.", "conduction_number"),
            Field("brand", "Brand", "brand"),
            Field("model", "Model", "model"),
            Field("year", "Year", "year", "number"),
            Field("variant", "Variant", "variant"),
            Field("color", "Colour", "color"),
            Field("engine_number", "Engine No.", "engine_number"),
            Field("chassis_number", "Chassis No.", "chassis_number"),
            Field("fuel_type", "Fuel Type", "fuel_type"),
            Field("status", "Status", "status"),
            Field("current_odometer", "Odometer", "current_odometer", "number"),
            Field("acquisition_date", "Acquisition Date", "acquisition_date", "date"),
            Field("acquisition_cost", "Acquisition Cost", "acquisition_cost", "number"),
            Field("far_number", "FAR No.", "far_number"),
            Field("branch_name", "Branch", "branch.name", join="branch"),
            Field("vehicle_type_name", "Vehicle Type", "vehicle_type.name",
                  join="vehicle_type"),
            Field("driver_name", "Assigned Driver", "assigned_driver.full_name",
                  join="assigned_driver"),
        ]),
    "maintenance_orders": DataSource(
        key="maintenance_orders", label="Maintenance Orders",
        model_path="app.modules.transactions.maintenance_order.models.MaintenanceOrder",
        permission="maintenanceorder.view",
        description="Work orders, with vehicle, maintenance type and vendor.",
        joins=[
            Join("vehicle", "vehicle", "Vehicle"),
            Join("maintenance_type", "maintenance_type", "Maintenance Type"),
            Join("transaction_type", "transaction_type", "Transaction Type"),
            Join("vendor", "vendor", "Vendor"),
        ],
        fields=[
            Field("document_number", "MO No.", "document_number"),
            Field("status", "Status", "status"),
            Field("order_category", "Category", "order_category"),
            Field("scheduled_date", "Scheduled", "scheduled_date", "date"),
            Field("completed_date", "Completed", "completed_date", "date"),
            Field("odometer_at_service", "Odometer", "odometer_at_service", "number"),
            Field("estimated_cost", "Est. Cost", "estimated_cost", "number"),
            Field("actual_cost", "Actual Cost", "actual_cost", "number"),
            Field("assigned_mechanic", "Mechanic", "assigned_mechanic"),
            Field("description", "Description", "description"),
            Field("plate_number", "Plate No.", "vehicle.plate_number",
                  join="vehicle"),
            Field("vehicle_brand", "Brand", "vehicle.brand", join="vehicle"),
            Field("vehicle_model", "Model", "vehicle.model", join="vehicle"),
            Field("maintenance_type_name", "Maintenance Type",
                  "maintenance_type.name", join="maintenance_type"),
            Field("transaction_type_name", "Transaction Type",
                  "transaction_type.name", join="transaction_type"),
            Field("vendor_name", "Vendor", "vendor.name", join="vendor"),
        ]),
    "vehicle_registrations": DataSource(
        key="vehicle_registrations", label="Vehicle Registrations",
        model_path="app.modules.transactions.vehicle_registration.models.VehicleRegistration",
        permission="vehicleregistration.view",
        description="LTO registration records with expiry dates.",
        joins=[Join("vehicle", "vehicle", "Vehicle")],
        fields=[
            Field("document_number", "Registration No.", "document_number"),
            Field("status", "Status", "status"),
            Field("or_number", "OR No.", "or_number"),
            Field("cr_number", "CR No.", "cr_number"),
            Field("expiry_date", "Expiry Date", "expiry_date", "date"),
            Field("plate_number", "Plate No.", "vehicle.plate_number",
                  join="vehicle"),
            Field("vehicle_brand", "Brand", "vehicle.brand", join="vehicle"),
        ]),
    "trip_tickets": DataSource(
        key="trip_tickets", label="Trip Tickets",
        model_path="app.modules.transactions.trip_ticket.models.TripTicket",
        permission="tripticket.view",
        description="Vehicle trips with destination and odometer readings.",
        joins=[Join("vehicle", "vehicle", "Vehicle")],
        fields=[
            Field("document_number", "Trip No.", "document_number"),
            Field("status", "Status", "status"),
            Field("trip_date", "Trip Date", "trip_date", "date"),
            Field("destination", "Destination", "destination"),
            Field("purpose", "Purpose", "purpose"),
            Field("odometer_out", "Odometer Out", "odometer_out", "number"),
            Field("odometer_in", "Odometer In", "odometer_in", "number"),
            Field("plate_number", "Plate No.", "vehicle.plate_number",
                  join="vehicle"),
        ]),
}

# Operator -> how it is applied. Every value is passed as a bound
# parameter; none of these build SQL by string concatenation.
OPERATORS = {
    "eq": "equals",
    "ne": "does not equal",
    "contains": "contains",
    "gt": "greater than",
    "gte": "greater than or equal to",
    "lt": "less than",
    "lte": "less than or equal to",
    "is_empty": "is empty",
    "not_empty": "is not empty",
}


class ReportBuilderError(Exception):
    pass


def _resolve(model, attr_path, joined):
    """Turn a dotted path into a real column, using the already-joined
    relationship rather than a fresh implicit join."""
    if "." not in attr_path:
        return getattr(model, attr_path)
    rel_name, column_name = attr_path.split(".", 1)
    related_model = joined.get(rel_name)
    if related_model is None:
        raise ReportBuilderError(
            f"Field '{attr_path}' needs the '{rel_name}' join, which was "
            f"not applied.")
    return getattr(related_model, column_name)


def _coerce(value, kind):
    if value in (None, ""):
        return None
    if kind == "number":
        return float(str(value).replace(",", ""))
    if kind == "date":
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(str(value), fmt).date()
            except ValueError:
                continue
        raise ReportBuilderError(f"'{value}' is not a valid date (use YYYY-MM-DD).")
    return str(value)


def run_report(source_key, field_keys, filters=None, sort_key=None,
               sort_dir="asc", limit=1000, user=None):
    """Execute a built report definition and return {columns, rows}.

    `user` is checked against the source's permission -- a saved report
    can never become a way around the normal access rules.
    """
    source = DATA_SOURCES.get(source_key)
    if source is None:
        raise ReportBuilderError(f"Unknown data source '{source_key}'.")
    if user is not None and not user.has_permission(source.permission):
        raise ReportBuilderError(
            f"You do not have permission to read {source.label} "
            f"(requires '{source.permission}').")

    field_keys = [k for k in (field_keys or []) if k in source.fields]
    if not field_keys:
        raise ReportBuilderError("Select at least one column.")
    fields = [source.fields[k] for k in field_keys]

    model = source.model
    query = db.session.query(model)

    # Only join what the chosen fields and filters actually need.
    needed = {f.join for f in fields if f.join}
    for flt in (filters or []):
        fld = source.fields.get(flt.get("field"))
        if fld and fld.join:
            needed.add(fld.join)

    joined = {}
    for join_key in needed:
        join = source.joins.get(join_key)
        if join is None:
            raise ReportBuilderError(f"Unknown join '{join_key}'.")
        relationship = getattr(model, join.relationship)
        query = query.outerjoin(relationship)
        joined[join.relationship] = relationship.property.mapper.class_

    conditions = []
    for flt in (filters or []):
        fld = source.fields.get(flt.get("field"))
        if fld is None:
            continue
        op = flt.get("op", "eq")
        if op not in OPERATORS:
            raise ReportBuilderError(f"Unknown operator '{op}'.")
        column = _resolve(model, fld.attr, joined)
        if op == "is_empty":
            conditions.append(column.is_(None))
            continue
        if op == "not_empty":
            conditions.append(column.isnot(None))
            continue
        value = _coerce(flt.get("value"), fld.kind)
        if value is None:
            continue
        if op == "eq":
            conditions.append(column == value)
        elif op == "ne":
            conditions.append(column != value)
        elif op == "contains":
            conditions.append(column.ilike(f"%{value}%"))
        elif op == "gt":
            conditions.append(column > value)
        elif op == "gte":
            conditions.append(column >= value)
        elif op == "lt":
            conditions.append(column < value)
        elif op == "lte":
            conditions.append(column <= value)
    if conditions:
        query = query.filter(and_(*conditions))

    if sort_key and sort_key in source.fields:
        sort_field = source.fields[sort_key]
        if sort_field.join and sort_field.join not in needed:
            join = source.joins[sort_field.join]
            relationship = getattr(model, join.relationship)
            query = query.outerjoin(relationship)
            joined[join.relationship] = relationship.property.mapper.class_
        column = _resolve(model, sort_field.attr, joined)
        query = query.order_by(column.desc() if sort_dir == "desc"
                              else column.asc())

    # Hard ceiling regardless of what was asked for -- a custom report
    # must never be able to pull the entire table into memory and take
    # the app down with it.
    limit = max(1, min(int(limit or 1000), 5000))
    records = query.limit(limit).all()

    rows = []
    for record in records:
        row = []
        for fld in fields:
            value = record
            for part in fld.attr.split("."):
                value = getattr(value, part, None)
                if value is None:
                    break
            row.append(value)
        rows.append(row)

    return {
        "columns": [{"key": f.key, "label": f.label, "kind": f.kind}
                   for f in fields],
        "rows": rows,
        "row_count": len(rows),
        "truncated": len(rows) >= limit,
    }


def to_xlsx(result, title="Custom Report") -> bytes:
    """Render a report result as an Excel workbook, for download or for
    emailing to another system."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Report"
    ws.append([c["label"] for c in result["columns"]])
    fill = PatternFill("solid", fgColor="1F3B4D")
    for idx in range(1, len(result["columns"]) + 1):
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        ws.column_dimensions[cell.column_letter].width = 20
    for row in result["rows"]:
        ws.append([
            v if isinstance(v, (int, float, date, datetime, type(None)))
            else str(v)
            for v in row])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
