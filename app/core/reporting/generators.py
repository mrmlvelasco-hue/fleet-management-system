"""Excel report generators.

Each function takes an optional `filters` dict and returns
(filename: str, xlsx_bytes: bytes). Kept separate from the Flask routes
so the exact same code path produces the file whether a person clicks
"Export to Excel" in the browser or a ScheduledReport emails it
unattended overnight — the two were previously at risk of drifting apart
since the manual export lived inline in a route.

Supported filter keys (all optional; a report ignores keys it doesn't use):
  branch_id       - int, restrict to one branch
  vehicle_type_id - int, restrict to one vehicle type
  plate_number    - str, case-insensitive substring match against plate
                    number OR conduction number
  date_from       - date, inclusive
  date_to         - date, inclusive
  status          - str, e.g. "OVERDUE" for the due-based reports
"""
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2",
                           fill_type="solid")
_SECTION_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2",
                            fill_type="solid")


def _vehicle_matches(vehicle, filters: dict) -> bool:
    """Shared per-vehicle filter check, mirroring the report pages' own
    _vehicle_matches_filters so the Excel export always contains exactly
    what the screen showed for the same filter values."""
    if filters.get("branch_id"):
        if vehicle.branch_id != int(filters["branch_id"]):
            return False
    if filters.get("vehicle_type_id"):
        if vehicle.vehicle_type_id != int(filters["vehicle_type_id"]):
            return False
    if filters.get("plate_number"):
        needle = str(filters["plate_number"]).lower()
        plate = (vehicle.plate_number or "").lower()
        conduction = (vehicle.conduction_number or "").lower()
        if needle not in plate and needle not in conduction:
            return False
    return True


def _new_workbook(title: str, sheet_name: str):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append([title])
    ws["A1"].font = Font(size=14, bold=True)
    ws.append([])
    return wb, ws


def _style_header_row(ws, row_num: int, ncols: int):
    for c in range(1, ncols + 1):
        ws.cell(row_num, c).font = Font(bold=True)
        ws.cell(row_num, c).fill = _HEADER_FILL


def _autosize(ws, headers: list):
    for i, label in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = max(
            12, len(str(label)) + 2)


def _to_bytes(wb) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Vehicle Register Details ────────────────────────────────────────────────

def generate_vehicle_register_xlsx(filters: dict = None, user=None):
    from app.modules.master_data.vehicle.report_service import (
        VehicleRegisterReportService)
    filters = filters or {}
    groups = VehicleRegisterReportService().get_grouped(user=user)
    if filters.get("branch_id"):
        groups = [g for g in groups
                 if str(g["branch_code"]) == str(filters["branch_id"])
                 or any(v.get("branch_code") == filters["branch_id"]
                       for v in g["vehicles"])]

    columns = [
        ("Plate No.", "plate_number"), ("Assignee", "assignee"),
        ("Make", "make"), ("Model", "model"), ("Variant", "variant"),
        ("Year", "year"), ("Displacement", "displacement"),
        ("Fuel", "fuel"), ("Transmission", "transmission"),
        ("BodyColor", "body_color"), ("Type", "type"),
        ("FARNumber", "far_number"), ("MVFileNo", "mv_file_number"),
        ("CRNumber", "cr_number"), ("EngineNumber", "engine_number"),
        ("ChassisNumber", "chassis_number"),
    ]
    wb, ws = _new_workbook("Vehicle Register Details", "Vehicle Register")
    for group in groups:
        ws.append([group["branch_code"]])
        r = ws.max_row
        ws.cell(r, 1).font = Font(bold=True)
        for c in range(1, len(columns) + 1):
            ws.cell(r, c).fill = _SECTION_FILL
        ws.append([label for label, _ in columns])
        _style_header_row(ws, ws.max_row, len(columns))
        for row in group["vehicles"]:
            ws.append([row[key] for _, key in columns])
        ws.append([])
    _autosize(ws, [c[0] for c in columns])
    return (f"Vehicle_Register_Details_{datetime.now():%Y%m%d}.xlsx",
           _to_bytes(wb))


# ── PMS Compliance / Due Report ─────────────────────────────────────────────

def generate_pms_compliance_xlsx(filters: dict = None, user=None):
    from app.core.maintenance.due_calculation_service import (
        PMDueCalculationService)
    filters = filters or {}
    rows = [r for r in PMDueCalculationService().get_all_due_vehicles()
           if _vehicle_matches(r["vehicle"], filters)]
    if filters.get("status"):
        rows = [r for r in rows if r["status"] == filters["status"]]

    columns = ["Plate No.", "Branch", "Make", "Model", "Maintenance Type",
              "Next Due (km)", "Current Odometer", "Next Due Date",
              "Status"]
    wb, ws = _new_workbook("PMS Compliance / Due Report", "PMS Compliance")
    ws.append(columns)
    _style_header_row(ws, ws.max_row, len(columns))
    for r in rows:
        v = r["vehicle"]
        ws.append([
            v.plate_number or v.conduction_number, v.branch.name if v.branch else "—",
            v.brand, v.model,
            r["schedule"].maintenance_type.name if r.get("schedule") and r["schedule"].maintenance_type else "—",
            r.get("next_due_km") or "—", v.current_odometer,
            r.get("next_due_date") or "—", r["status"],
        ])
    _autosize(ws, columns)
    return (f"PMS_Compliance_Report_{datetime.now():%Y%m%d}.xlsx",
           _to_bytes(wb))


# ── Vehicle Registration Expiry Report ──────────────────────────────────────

def generate_registration_expiry_xlsx(filters: dict = None, user=None):
    from app.modules.registration_config.service import (
        RegistrationDueCalculationService)
    filters = filters or {}
    all_statuses = ("OVERDUE", "DUE_SOON", "GOOD", "NO_RECORD")
    rows = RegistrationDueCalculationService().get_all_due_vehicles(
        statuses=all_statuses)
    rows = [r for r in rows if _vehicle_matches(r["vehicle"], filters)]
    if filters.get("status"):
        rows = [r for r in rows if r["status"] == filters["status"]]

    columns = ["Plate No.", "Branch", "Make", "Model", "LTO Month",
              "LTO Week", "Next Due Date", "Source", "Status", "Warning"]
    wb, ws = _new_workbook("Vehicle Registration Expiry Report",
                           "Registration Expiry")
    ws.append(columns)
    _style_header_row(ws, ws.max_row, len(columns))
    for r in rows:
        v = r["vehicle"]
        ws.append([
            v.plate_number or v.conduction_number, v.branch.name if v.branch else "—",
            v.brand, v.model, r.get("lto_month") or "—", r.get("lto_week") or "—",
            r.get("next_due_date") or "—", r.get("source") or "—", r["status"],
            r.get("warning") or "",
        ])
    _autosize(ws, columns)
    return (f"Registration_Expiry_Report_{datetime.now():%Y%m%d}.xlsx",
           _to_bytes(wb))


# ── Maintenance Cost Summary ─────────────────────────────────────────────────

def generate_maintenance_cost_summary_xlsx(filters: dict = None, user=None):
    from app.extensions import db
    from app.modules.transactions.maintenance_order.models import (
        MaintenanceOrder)
    from app.modules.user_management.org_scope_service import (
        UserOrgScopeService)

    filters = filters or {}
    query = MaintenanceOrder.query.filter_by(status="COMPLETED")
    if filters.get("branch_id"):
        query = query.filter(MaintenanceOrder.vehicle.has(
            branch_id=int(filters["branch_id"])))
    if filters.get("vehicle_type_id"):
        query = query.filter(MaintenanceOrder.vehicle.has(
            vehicle_type_id=int(filters["vehicle_type_id"])))
    if filters.get("date_from"):
        query = query.filter(MaintenanceOrder.completed_date >= filters["date_from"])
    if filters.get("date_to"):
        query = query.filter(MaintenanceOrder.completed_date <= filters["date_to"])

    orders = query.order_by(MaintenanceOrder.completed_date.desc()).all()

    if filters.get("plate_number"):
        needle = str(filters["plate_number"]).lower()
        orders = [o for o in orders
                 if needle in (o.vehicle.plate_number or "").lower()
                 or needle in (o.vehicle.conduction_number or "").lower()]

    if user is not None:
        scope_svc = UserOrgScopeService()
        orders = [o for o in orders
                 if scope_svc.covers(user.id, branch_id=o.vehicle.branch_id)]

    columns = ["MO Number", "Vehicle", "Branch", "Category",
              "Maintenance Type", "Completed Date", "Actual Cost"]
    wb, ws = _new_workbook("Maintenance Cost Summary", "Cost Summary")
    ws.append(columns)
    _style_header_row(ws, ws.max_row, len(columns))
    total = 0
    for o in orders:
        cost = float(o.actual_cost or 0)
        total += cost
        ws.append([
            o.document_number or "(draft)",
            f"{o.vehicle.plate_number or o.vehicle.conduction_number} — {o.vehicle.brand} {o.vehicle.model}",
            o.vehicle.branch.name if o.vehicle.branch else "—",
            o.category or o.order_category,
            o.maintenance_type.name if o.maintenance_type else "—",
            o.completed_date, cost,
        ])
    ws.append([])
    ws.append(["", "", "", "", "", "TOTAL", total])
    ws.cell(ws.max_row, 6).font = Font(bold=True)
    ws.cell(ws.max_row, 7).font = Font(bold=True)
    _autosize(ws, columns)
    return (f"Maintenance_Cost_Summary_{datetime.now():%Y%m%d}.xlsx",
           _to_bytes(wb))


def generate_audit_trail_xlsx(filters: dict = None, user=None):
    """Filters: table, action, user_id, date_from, date_to -- same keys
    the Audit Trail page's own query string uses, so Export always
    reflects exactly what's on screen."""
    from app.core.models.audit_log import AuditLog
    from app.modules.user_management.models import User
    from app.extensions import db
    filters = filters or {}

    q = AuditLog.query
    if filters.get("table"):
        q = q.filter(AuditLog.table_name == filters["table"])
    if filters.get("action"):
        q = q.filter(AuditLog.action == filters["action"])
    if filters.get("user_id"):
        q = q.filter(AuditLog.user_id == int(filters["user_id"]))
    if filters.get("date_from"):
        q = q.filter(AuditLog.timestamp >= filters["date_from"])
    if filters.get("date_to"):
        q = q.filter(AuditLog.timestamp <= filters["date_to"] + " 23:59:59")
    logs = q.order_by(AuditLog.id.desc()).limit(5000).all()

    columns = ["ID", "Timestamp", "Table", "Record ID", "Action", "User",
              "Changes"]
    wb, ws = _new_workbook("Audit Trail", "Audit Trail")
    ws.append(columns)
    _style_header_row(ws, ws.max_row, len(columns))
    for log in logs:
        user_obj = db.session.get(User, log.user_id) if log.user_id else None
        user_label = (user_obj.username if user_obj else
                     (str(log.user_id) if log.user_id else "—"))
        changes = ""
        if log.action == "UPDATE" and log.old_values and log.new_values:
            diffs = []
            for k, new_v in (log.new_values or {}).items():
                old_v = (log.old_values or {}).get(k)
                if old_v != new_v:
                    diffs.append(f"{k}: {old_v} -> {new_v}")
            changes = "; ".join(diffs)
        ws.append([log.id, log.timestamp, log.table_name, log.record_id,
                  log.action, user_label, changes])
    _autosize(ws, columns)
    return (f"Audit_Trail_{datetime.now():%Y%m%d}.xlsx", _to_bytes(wb))


def generate_vehicle_activity_history_xlsx(vehicle_ids: list, user=None):
    """One sheet per selected vehicle: Vehicle Master Information, the
    full Activity History timeline, the Utilization Summary, and the
    Outlet Assignment History. Deliberately NOT part of REPORT_GENERATORS/
    the scheduled-report emailer -- this report's "filter" is a specific
    list of vehicle ids picked by the person running it, which doesn't
    fit the scheduler's generic branch/status/date filters model (there's
    no sensible recurring schedule for "always email me these 3 specific
    vehicles")."""
    from app.core.vehicle_activity_history_service import (
        VehicleActivityHistoryService)
    from app.modules.master_data.vehicle.models import Vehicle

    activity_svc = VehicleActivityHistoryService()
    wb = Workbook()
    wb.remove(wb.active)

    for vid in vehicle_ids:
        from app.extensions import db as _db
        vehicle = _db.session.get(Vehicle, vid)
        if vehicle is None:
            continue
        sheet_name = (vehicle.plate_number or vehicle.conduction_number
                     or f"Vehicle {vid}")[:31]  # Excel sheet name limit
        ws = wb.create_sheet(title=sheet_name)

        ws.append(["Vehicle Master Information"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=13)
        ws.append(["Plate No.", vehicle.plate_number or vehicle.conduction_number,
                   "Vehicle", f"{vehicle.brand} {vehicle.model} {vehicle.variant or ''}"])
        ws.append(["Year Model", vehicle.year, "Engine No.", vehicle.engine_number or "—"])
        ws.append(["Chassis No.", vehicle.chassis_number or "—",
                   "Acquisition Cost", float(vehicle.acquisition_cost) if vehicle.acquisition_cost else None])
        ws.append(["Current Status", vehicle.status])
        ws.append([])

        rows = activity_svc.get_activity_rows(vehicle)
        ws.append(["Vehicle Activity History"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
        header_row = ws.max_row + 1
        ws.append(["Date", "Activity Type", "Outlet/Department", "Assigned To",
                  "Description", "Cost", "Odometer"])
        _style_header_row(ws, header_row, 7)
        for r in rows:
            ws.append([r["date"], r["activity_type"], r["outlet"],
                      r["assigned_to"] or "—", r["description"],
                      float(r["cost"]) if r["cost"] is not None else None,
                      r["odometer"]])
        ws.append([])

        util = activity_svc.get_utilization_summary(vehicle, rows)
        ws.append(["Vehicle Utilization Summary"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
        ws.append(["Total Transfers", util["total_transfers"]])
        ws.append(["PMS Activities", util["pms_count"]])
        ws.append(["Repair Activities", util["repair_count"]])
        ws.append(["Tire Replacements", util["tire_replacements"]])
        ws.append(["Battery Replacements", util["battery_replacements"]])
        ws.append(["Total Maintenance Cost", float(util["total_maintenance_cost"])])
        ws.append(["No. of Assigned Outlets", util["assigned_outlets_count"]])
        ws.append(["Vehicle Age (Years)", util["vehicle_age_years"]])
        ws.append(["Current Odometer", util["current_odometer"]])
        ws.append([])

        outlet_history = activity_svc.get_outlet_history(vehicle)
        ws.append(["Outlet Assignment History"])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
        header_row2 = ws.max_row + 1
        ws.append(["From Date", "To Date", "Outlet", "Custodian"])
        _style_header_row(ws, header_row2, 4)
        for seg in outlet_history:
            ws.append([seg["from_date"], seg["to_date"] or "Present",
                      seg["outlet"], seg["custodian"]])

        for col in "ABCDEFG":
            ws.column_dimensions[col].width = 20

    if not wb.sheetnames:
        wb.create_sheet(title="No Vehicles Selected")

    return (f"Vehicle_Activity_History_{datetime.now():%Y%m%d}.xlsx",
           _to_bytes(wb))


# Registry used by both manual export routes and the scheduled emailer.
REPORT_GENERATORS = {
    "RPT_VEHICLE_REGISTER": generate_vehicle_register_xlsx,
    "RPT_PMS_COMPLIANCE": generate_pms_compliance_xlsx,
    "RPT_REGISTRATION_EXPIRY": generate_registration_expiry_xlsx,
    "RPT_MAINTENANCE_COST_SUMMARY": generate_maintenance_cost_summary_xlsx,
    "RPT_AUDIT_TRAIL": generate_audit_trail_xlsx,
}
