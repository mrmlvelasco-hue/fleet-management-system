"""Vehicle bulk import from a fill-in Excel workbook.

Two halves:
  * build_template()  -> a blank .xlsx with the exact expected headers,
    an example row, and a reference sheet listing the valid codes for
    every lookup column, so a person can fill it in without guessing.
  * import_vehicles()  -> validates and loads a filled-in workbook,
    reporting per-row errors rather than failing the whole file on one
    bad cell.

Design notes:
  * Lookup columns are matched by CODE (branch code, vehicle type code),
    not database id -- an id is meaningless to whoever fills in the
    sheet, and codes are stable and visible in the app's own screens.
  * dry_run=True validates everything and writes nothing, so a person
    can see exactly what would happen before committing. This mirrors
    the PM import script's behaviour.
  * A row that fails validation is SKIPPED and reported; the remaining
    rows still import. A partial import is far more useful for a
    migration than an all-or-nothing failure on row 200 of 400.
"""
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.extensions import db

# (column header, required?, human-readable note for the reference sheet)
TEMPLATE_COLUMNS = [
    ("conduction_number", False, "Conduction sticker no. Required if Plate No. is blank."),
    ("plate_number", False, "LTO plate no. Required if Conduction No. is blank."),
    ("vehicle_type_code", True, "Must match a Vehicle Type code (see 'Reference' sheet)."),
    ("branch_code", True, "Must match a Branch code (see 'Reference' sheet)."),
    ("brand", True, "e.g. Mitsubishi"),
    ("model", True, "e.g. Strada"),
    ("year", True, "4-digit year, e.g. 2013"),
    ("variant", False, "e.g. GL, 4x2"),
    ("color", False, "e.g. White"),
    ("engine_number", False, "Must be unique if provided."),
    ("chassis_number", False, "Must be unique if provided."),
    ("fuel_type", False, "e.g. DIESEL, GASOLINE"),
    ("transmission", False, "e.g. MANUAL, AUTOMATIC"),
    ("current_odometer", False, "Whole km, e.g. 60000"),
    ("acquisition_date", False, "YYYY-MM-DD"),
    ("acquisition_cost", False, "Numbers only, e.g. 950000"),
    ("far_number", False, "Fixed Asset Register no."),
    ("cr_number", False, "Certificate of Registration no."),
    ("status", False, "ACTIVE (default), IN_REPAIR, INACTIVE"),
]

REQUIRED_COLUMNS = [c for c, required, _ in TEMPLATE_COLUMNS if required]


def build_template() -> bytes:
    """A blank workbook with the expected headers, one example row, and
    a Reference sheet of valid codes."""
    from app.modules.master_data.reference.models import VehicleType
    from app.modules.master_data.org.models import Branch

    wb = Workbook()
    ws = wb.active
    ws.title = "Vehicles"

    headers = [c for c, _, _ in TEMPLATE_COLUMNS]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1F3B4D")
    for idx, (name, required, _note) in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = max(16, len(name) + 4)

    # One example row so the expected formats are unambiguous. Clearly
    # marked so it isn't mistaken for real data.
    ws.append(["EXAMPLE-001", "ABC1234", "LV", "MNL", "Mitsubishi",
               "Strada", 2013, "GL", "White", "4D56UCEM5302",
               "MMBJNKA40ED011014", "DIESEL", "MANUAL", 60000,
               "2013-05-15", 950000, "FAR-0001", "CR-0001", "ACTIVE"])
    for idx in range(1, len(headers) + 1):
        ws.cell(row=2, column=idx).font = Font(italic=True, color="888888")
    ws.cell(row=3, column=1).value = "^ Delete this example row before uploading."
    ws.cell(row=3, column=1).font = Font(italic=True, bold=True, color="C00000")

    ref = wb.create_sheet("Reference")
    ref.append(["Column", "Required", "Notes"])
    for c in ref[1]:
        c.font = Font(bold=True)
    for name, required, note in TEMPLATE_COLUMNS:
        ref.append([name, "YES" if required else "optional", note])
    ref.column_dimensions["A"].width = 24
    ref.column_dimensions["B"].width = 12
    ref.column_dimensions["C"].width = 70

    ref.append([])
    ref.append(["Valid Vehicle Type codes:"])
    ref.cell(ref.max_row, 1).font = Font(bold=True)
    for vt in VehicleType.query.filter_by(is_active=True).all():
        ref.append([vt.code, vt.name])

    ref.append([])
    ref.append(["Valid Branch codes:"])
    ref.cell(ref.max_row, 1).font = Font(bold=True)
    for b in Branch.query.filter_by(is_active=True).all():
        ref.append([b.code, b.name])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _clean(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def import_vehicles(file_stream, dry_run: bool = True) -> dict:
    """Validate (and optionally load) a filled-in template.

    Returns a summary with per-row errors so the person can fix the
    sheet and re-upload, rather than being told only that "the import
    failed".
    """
    from app.modules.master_data.reference.models import VehicleType
    from app.modules.master_data.org.models import Branch
    from app.modules.master_data.vehicle.models import Vehicle
    from app.modules.master_data.vehicle.service import (
        VehicleService, DuplicateVehicleError)

    wb = load_workbook(file_stream, data_only=True)
    ws = wb["Vehicles"] if "Vehicles" in wb.sheetnames else wb[wb.sheetnames[0]]

    header_row = [_clean(c.value) for c in ws[1]]
    missing = [c for c in REQUIRED_COLUMNS if c not in header_row]
    if missing:
        raise ValueError(
            f"The uploaded file is missing required column(s): "
            f"{', '.join(missing)}. Please start from the downloaded "
            f"template.")
    idx = {name: i for i, name in enumerate(header_row) if name}

    branches = {b.code.upper(): b.id
                for b in Branch.query.all() if b.code}
    vtypes = {v.code.upper(): v.id
              for v in VehicleType.query.all() if v.code}

    stats = {"total_rows": 0, "created": 0, "skipped": 0, "errors": []}
    svc = VehicleService()

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True),
                                     start=2):
        if row is None or all(_clean(v) is None for v in row):
            continue  # blank spacer row
        get = lambda col: (_clean(row[idx[col]])
                          if col in idx and idx[col] < len(row) else None)

        conduction = get("conduction_number")
        plate = get("plate_number")
        # Skip the template's own example row rather than trying to
        # import it -- people routinely forget to delete it.
        if conduction == "EXAMPLE-001":
            continue
        if (conduction or "").startswith("^ Delete this example"):
            continue

        stats["total_rows"] += 1
        errors = []

        if not conduction and not plate:
            errors.append("needs a conduction_number or a plate_number")

        vt_code = (get("vehicle_type_code") or "").upper()
        if not vt_code:
            errors.append("vehicle_type_code is required")
        elif vt_code not in vtypes:
            errors.append(f"unknown vehicle_type_code '{vt_code}'")

        br_code = (get("branch_code") or "").upper()
        if not br_code:
            errors.append("branch_code is required")
        elif br_code not in branches:
            errors.append(f"unknown branch_code '{br_code}'")

        brand, model, year = get("brand"), get("model"), get("year")
        if not brand:
            errors.append("brand is required")
        if not model:
            errors.append("model is required")
        year_value = None
        if not year:
            errors.append("year is required")
        else:
            try:
                year_value = int(float(year))
            except (TypeError, ValueError):
                errors.append(f"year '{year}' is not a number")

        if conduction and Vehicle.query.filter_by(
                conduction_number=conduction).first():
            errors.append(f"conduction_number '{conduction}' already exists")
        if plate and Vehicle.query.filter_by(plate_number=plate).first():
            errors.append(f"plate_number '{plate}' already exists")

        if errors:
            stats["skipped"] += 1
            stats["errors"].append({"row": row_number,
                                   "identifier": plate or conduction or "(blank)",
                                   "problems": errors})
            continue

        if dry_run:
            stats["created"] += 1
            continue

        optional = {}
        for col, caster in (("variant", str), ("color", str),
                           ("engine_number", str), ("chassis_number", str),
                           ("fuel_type", str), ("transmission", str),
                           ("far_number", str), ("cr_number", str),
                           ("status", str)):
            value = get(col)
            if value:
                optional[col] = caster(value)
        for col in ("current_odometer",):
            value = get(col)
            if value:
                try:
                    optional[col] = int(float(str(value).replace(",", "")))
                except (TypeError, ValueError):
                    pass
        for col in ("acquisition_cost",):
            value = get(col)
            if value:
                try:
                    optional[col] = float(str(value).replace(",", ""))
                except (TypeError, ValueError):
                    pass
        acq = get("acquisition_date")
        if acq:
            from datetime import datetime as _dt
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
                try:
                    optional["acquisition_date"] = _dt.strptime(acq, fmt).date()
                    break
                except ValueError:
                    continue

        try:
            svc.create(vehicle_type_id=vtypes[vt_code],
                       brand=brand, model=model, year=year_value,
                       branch_id=branches[br_code],
                       conduction_number=conduction, plate_number=plate,
                       **optional)
            stats["created"] += 1
        except DuplicateVehicleError as exc:
            stats["skipped"] += 1
            stats["errors"].append({"row": row_number,
                                   "identifier": plate or conduction,
                                   "problems": [str(exc)]})
        except Exception as exc:  # one bad row must not kill the batch
            db.session.rollback()
            stats["skipped"] += 1
            stats["errors"].append({"row": row_number,
                                   "identifier": plate or conduction,
                                   "problems": [f"unexpected error: {exc}"]})

    return stats
